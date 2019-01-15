# -*- coding: utf-8 -*-

"""
假定你一直“自愿”为“强制自愿俱乐部”记录会员会费。这确实是一项枯燥
的工作，包括维护一个电子表格，记录每个月谁交了会费，并用电子邮件提醒那些
没交的会员。不必你自己查看电子表格，而是向会费超期的会员复制和粘贴相同的
电子邮件。
"""

import openpyxl, smtplib, sys
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

print('input your gmail emailAddress')
emailAddress = input()
print('input your password')
password = input()
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(emailAddress, password)

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues " \
           "for % s.Please make this payment as soon as possible.Thank you!'" % (latestMonth, name, latestMonth)

    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
    smtpObj.quit()