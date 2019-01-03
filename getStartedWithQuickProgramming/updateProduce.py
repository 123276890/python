# -*- coding: utf-8 -*-

"""
这个项目需要编写一个程序，更新产品销售电子表格中的单元格。程序将遍
历这个电子表格，找到特定类型的产品，并更新它们的价格
"""

import openpyxl, os

os.chdir('/Users/zhuangganglong/Downloads/automate_online-materials')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

price_updates = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]

os.chdir('/Users/zhuangganglong/python/getStartedWithQuickProgramming')
wb.save('updateProduceSales.xlsx')